from flask import Flask, render_template, redirect, url_for, request, flash, jsonify, send_file
from flask_caching import Cache
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from models import db, User, Category, Wallet, Transaction, Budget, SharedWallet
from collections import defaultdict
from datetime import datetime
from dotenv import load_dotenv
import os
from urllib.parse import urlencode
from sqlalchemy.orm import joinedload
from utils.datetime_utils import WIB, now_wib, to_wib
from utils.logger import setup_logger
from api import api_bp
from services.transaction_service import (
    calculate_transaction_totals,
    create_transaction,
    delete_transaction as delete_transaction_service,
    get_filtered_transactions,
    normalize_wib_storage,
    parse_positive_amount,
    parse_transaction_datetime,
    update_transaction,
)
from services.wallet_service import (
    create_wallet,
    delete_wallet as delete_wallet_service,
    transfer_balance,
    update_wallet,
)

# load environment variables from .env file (if present)
load_dotenv()
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape, portrait
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from PIL import Image, ImageOps, UnidentifiedImageError

app = Flask(__name__)
logger = setup_logger()
# gunakan environment variable untuk secret key agar tidak tersimpan di kode
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'kunci-rahasia-ubah-sekarang')
# durasi cookie untuk fitur "ingat saya" (opsional, default 365 hari)
from datetime import timedelta
app.config['REMEMBER_COOKIE_DURATION'] = timedelta(days=30)

BASE_DIR = os.path.abspath(os.path.dirname(__file__))


def resolve_database_path():
    db_path = os.environ.get('DB_PATH', 'dev.db').strip()
    if not db_path:
        db_path = 'dev.db'

    if not os.path.isabs(db_path):
        db_path = os.path.join(BASE_DIR, db_path)

    return os.path.abspath(db_path)


db_path = resolve_database_path()

app.config['SQLALCHEMY_DATABASE_URI'] = f"sqlite:///{db_path.replace(os.sep, '/')}"
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = os.path.join(app.root_path, 'static', 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['CACHE_TYPE'] = 'simple'
app.config['CACHE_DEFAULT_TIMEOUT'] = 60

ALLOWED_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
MAX_PROFILE_IMAGE_DIMENSION = 720
MAX_PROFILE_IMAGE_PIXELS = 20_000_000
PROFILE_JPEG_QUALITY = 78

db.init_app(app)
cache = Cache(config={"CACHE_TYPE": "simple"})
cache.init_app(app)
app.register_blueprint(api_bp)

from sqlalchemy import event
from sqlalchemy.engine import Engine
import sqlite3

@event.listens_for(Engine, "connect")
def set_sqlite_pragma(dbapi_connection, connection_record):
    if not app.config['SQLALCHEMY_DATABASE_URI'].startswith('sqlite:///') or not isinstance(dbapi_connection, sqlite3.Connection):
        return

    cursor = dbapi_connection.cursor()
    cursor.execute("PRAGMA journal_mode=DELETE;")
    cursor.execute("PRAGMA synchronous=NORMAL")
    cursor.execute("PRAGMA cache_size=10000")
    cursor.execute("PRAGMA temp_store=MEMORY")
    cursor.close()

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


@cache.memoize(timeout=60)
def get_dashboard_data(user_id, selected_month=None, trend_year=None):
    now = normalize_wib_storage(now_wib())

    total_balance = db.session.query(db.func.sum(Wallet.balance)).filter_by(user_id=user_id).scalar() or 0

    if not selected_month:
        selected_month = now.strftime('%Y-%m')

    chart_query = Transaction.query.options(
        joinedload(Transaction.category),
    ).filter_by(user_id=user_id)

    if selected_month:
        try:
            year, month = map(int, selected_month.split('-'))
            chart_query = chart_query.filter(
                db.extract('year', Transaction.date) == year,
                db.extract('month', Transaction.date) == month,
            )
        except (TypeError, ValueError):
            selected_month = now.strftime('%Y-%m')
            year, month = map(int, selected_month.split('-'))
            chart_query = chart_query.filter(
                db.extract('year', Transaction.date) == year,
                db.extract('month', Transaction.date) == month,
            )

    transactions = chart_query.order_by(Transaction.date.desc()).all()
    totals = calculate_transaction_totals(transactions)

    expense_by_category = {}
    for transaction in transactions:
        if transaction.type != 'expense':
            continue

        category_name = transaction.category.name if transaction.category else 'Tanpa Kategori'
        expense_by_category[category_name] = expense_by_category.get(category_name, 0.0) + float(transaction.amount or 0)

    if not trend_year:
        trend_year = now.year

    trend_transactions = Transaction.query.filter_by(user_id=user_id).filter(
        db.extract('year', Transaction.date) == trend_year
    ).all()

    income_map = defaultdict(float)
    expense_map = defaultdict(float)
    for transaction in trend_transactions:
        month_key = transaction.date.strftime('%m')
        if transaction.type == 'income':
            income_map[month_key] += float(transaction.amount or 0)
        else:
            expense_map[month_key] += float(transaction.amount or 0)

    trend_labels = ['Jan', 'Feb', 'Mar', 'Apr', 'Mei', 'Jun', 'Jul', 'Agu', 'Sep', 'Okt', 'Nov', 'Des']
    month_keys = [str(i).zfill(2) for i in range(1, 13)]
    income_data = [income_map[month] for month in month_keys]
    expense_data = [expense_map[month] for month in month_keys]

    recent_transactions = Transaction.query.options(
        joinedload(Transaction.category),
        joinedload(Transaction.wallet),
    ).filter_by(user_id=user_id).order_by(Transaction.date.desc()).limit(5).all()

    return {
        'total_balance': total_balance,
        'chart_income': totals['total_income'],
        'chart_expense': totals['total_expense'],
        'chart_category_labels': list(expense_by_category.keys()),
        'chart_category_values': list(expense_by_category.values()),
        'recent_transactions': recent_transactions,
        'selected_month': selected_month,
        'trend_year': trend_year,
        'trend_labels': trend_labels,
        'trend_income_data': income_data,
        'trend_expense_data': expense_data,
    }

@app.context_processor
def inject_categories_wallets():
    try:
        if current_user.is_authenticated:
            categories = Category.query.filter_by(user_id=current_user.id).all()
            wallets = Wallet.query.filter_by(user_id=current_user.id).all()
            shared_wallets = SharedWallet.query.filter_by(shared_with_id=current_user.id, permission='add').all()
            shared_wallet_objects = [sw.wallet for sw in shared_wallets]
            all_wallets = wallets + shared_wallet_objects
            return dict(
                categories=categories,
                wallets=all_wallets,
                datetime=datetime,
                to_wib=to_wib,
                now_wib=now_wib,
                profile_photo_url=get_profile_photo_url(current_user.photo)
            )
    except Exception:
        pass
    return dict(categories=[], wallets=[], datetime=datetime, to_wib=to_wib, now_wib=now_wib, profile_photo_url=None)

# Buat direktori instance jika belum ada
os.makedirs(os.path.join(app.root_path, 'instance'), exist_ok=True)

# Buat direktori untuk upload foto profil
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

with app.app_context():
    db.create_all()

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        # Cek apakah username sudah ada
        existing_user = User.query.filter_by(username=username).first()
        if existing_user:
            flash('Username sudah digunakan')
            return redirect(url_for('register'))
        user = User(username=username)
        user.set_password(password)  # Hash password sebelum menyimpan
        db.session.add(user)
        db.session.flush()  # Dapatkan id user sebelum commit

        # Buat kategori default
        default_categories = [
            Category(name='Gaji', type='income', user_id=user.id),
            Category(name='Hadiah', type='income', user_id=user.id),
            Category(name='Makanan', type='expense', user_id=user.id),
            Category(name='Transport', type='expense', user_id=user.id),
            Category(name='Belanja', type='expense', user_id=user.id),
            Category(name='Tagihan', type='expense', user_id=user.id),
            Category(name='Transfer (Keluar)', type='expense', user_id=user.id),
            Category(name='Transfer (Masuk)', type='income', user_id=user.id),
            Category(name='Biaya Transfer', type='expense', user_id=user.id),
        ]
        db.session.add_all(default_categories)
        db.session.commit()
        flash('Registrasi berhasil, silakan login')
        return redirect(url_for('login'))
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        user = User.query.filter_by(username=request.form['username']).first()
        if user and user.check_password(request.form['password']):
            # cek apakah user memilih "ingat saya"
            remember = True if request.form.get('remember') else False
            login_user(user, remember=remember)
            return redirect(url_for('dashboard'))
        flash('Username atau password salah')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/offline')
def offline():
    return render_template('offline.html')

@app.route('/')
@app.route('/dashboard')
@login_required
def dashboard():
    selected_month = request.args.get('month')
    trend_year = request.args.get('trend_year', type=int)

    if not selected_month:
        selected_month = normalize_wib_storage(now_wib()).strftime('%Y-%m')

    if not trend_year:
        trend_year = normalize_wib_storage(now_wib()).year

    dashboard_data = get_dashboard_data(current_user.id, selected_month, trend_year)

    return render_template('index.html', 
                           total_balance=dashboard_data['total_balance'],
                           income=dashboard_data['chart_income'],
                           expense=dashboard_data['chart_expense'],
                           category_labels=dashboard_data['chart_category_labels'],
                           category_values=dashboard_data['chart_category_values'],
                           recent_transactions=dashboard_data['recent_transactions'],
                           selected_month=dashboard_data['selected_month'],
                           trend_year=dashboard_data['trend_year'],
                           trend_labels=dashboard_data['trend_labels'],
                           trend_income=dashboard_data['trend_income_data'],
                           trend_expense=dashboard_data['trend_expense_data'])

@app.route('/categories')
@login_required
def categories():
    cats = Category.query.filter_by(user_id=current_user.id).all()
    return render_template('categories.html', categories=cats)

@app.route('/category/add', methods=['POST'])
@login_required
def add_category():
    name = request.form['name']
    type_ = request.form['type']
    cat = Category(name=name, type=type_, user_id=current_user.id)
    db.session.add(cat)
    db.session.commit()
    flash('Kategori ditambahkan')
    return redirect(url_for('categories'))

@app.route('/category/edit/<int:id>', methods=['POST'])
@login_required
def edit_category(id):
    cat = Category.query.get_or_404(id)
    if cat.user_id != current_user.id:
        flash('Anda tidak memiliki akses')
        return redirect(url_for('categories'))
    cat.name = request.form['name']
    cat.type = request.form['type']
    db.session.commit()
    flash('Kategori diperbarui')
    return redirect(url_for('categories'))

@app.route('/category/delete/<int:id>')
@login_required
def delete_category(id):
    cat = Category.query.get_or_404(id)
    if cat.user_id != current_user.id:
        flash('Anda tidak memiliki akses')
        return redirect(url_for('categories'))
    db.session.delete(cat)
    db.session.commit()
    flash('Kategori dihapus')
    return redirect(url_for('categories'))

@app.route('/wallets')
@login_required
def wallets():
    owned = Wallet.query.filter_by(user_id=current_user.id).all()
    # Dompet yang dibagikan kepada saya
    shared = SharedWallet.query.filter_by(shared_with_id=current_user.id).all()
    return render_template('wallets.html', owned=owned, shared=shared)

@app.route('/wallet/add', methods=['POST'])
@login_required
def add_wallet():
    name = request.form['name']
    type_ = request.form['type']
    try:
        balance = parse_positive_amount(request.form['balance'], 'Saldo awal', allow_zero=True)
    except ValueError as exc:
        flash(str(exc), 'danger')
        return redirect(url_for('wallets'))
    try:
        create_wallet(current_user.id, name, type_, balance)
        logger.info(f"Wallet created: user={current_user.id}, name={name}")
        flash('Dompet ditambahkan')
    except Exception:
        logger.error('Wallet create error', exc_info=True)
        flash('Gagal menambahkan dompet', 'danger')
    return redirect(url_for('wallets'))

@app.route('/wallet/edit/<int:id>', methods=['POST'])
@login_required
def edit_wallet(id):
    wallet = Wallet.query.get_or_404(id)
    try:
        balance = parse_positive_amount(request.form['balance'], 'Saldo', allow_zero=True)
    except ValueError as exc:
        flash(str(exc), 'danger')
        return redirect(url_for('wallets'))

    try:
        update_wallet(wallet, current_user.id, request.form['name'], request.form['type'], balance)
        logger.info(f"Wallet updated: user={current_user.id}, wallet_id={wallet.id}")
        flash('Dompet diperbarui')
    except ValueError as exc:
        flash(str(exc), 'danger')
    except Exception:
        logger.error('Wallet update error', exc_info=True)
        flash('Gagal memperbarui dompet', 'danger')
    return redirect(url_for('wallets'))

@app.route('/wallet/delete/<int:id>')
@login_required
def delete_wallet(id):
    wallet = Wallet.query.get_or_404(id)
    try:
        delete_wallet_service(wallet, current_user.id)
        logger.info(f"Wallet deleted: user={current_user.id}, wallet_id={id}")
        flash('Dompet dihapus')
    except ValueError as exc:
        flash(str(exc), 'danger')
    except Exception:
        logger.error('Wallet delete error', exc_info=True)
        flash('Gagal menghapus dompet', 'danger')
    return redirect(url_for('wallets'))

@app.route('/transactions')
@login_required
def transactions():
    # filter parameters
    category_filter = request.args.get('category_id', type=int)
    wallet_filter = request.args.get('wallet_id', type=int)
    start_date = request.args.get('start_date')  # expected YYYY-MM-DD
    end_date = request.args.get('end_date')      # expected YYYY-MM-DD
    search = request.args.get('search', '')
    page = max(request.args.get('page', 1, type=int), 1)
    per_page = request.args.get('per_page', 10, type=int)
    per_page = min(max(per_page, 1), 100)

    filters = {
        'category_id': category_filter,
        'wallet_id': wallet_filter,
        'start_date': start_date,
        'end_date': end_date,
        'search': search,
    }

    query = get_filtered_transactions(current_user.id, filters)
    trans = query.order_by(Transaction.date.desc()).paginate(
        page=page,
        per_page=per_page,
        error_out=False,
    )
    categories = Category.query.filter_by(user_id=current_user.id).all()
    wallets = Wallet.query.filter_by(user_id=current_user.id).all()
    # Tambahkan dompet bersama yang memiliki izin 'add'
    shared_wallets = SharedWallet.query.filter_by(shared_with_id=current_user.id, permission='add').all()
    shared_wallet_objects = [sw.wallet for sw in shared_wallets]
    all_wallets = wallets + shared_wallet_objects
    pagination_params = {
        'category_id': category_filter,
        'wallet_id': wallet_filter,
        'start_date': start_date,
        'end_date': end_date,
        'search': search,
        'per_page': per_page,
    }
    pagination_query = urlencode({
        key: value for key, value in pagination_params.items()
        if value not in (None, '')
    })
    return render_template('transactions.html', transactions=trans, categories=categories, wallets=all_wallets,
                           category_filter=category_filter, wallet_filter=wallet_filter,
                           start_date=start_date, end_date=end_date, search=search,
                           per_page=per_page, pagination_query=pagination_query)


@app.route('/report/preview')
@login_required
def report_preview():
    filters = {
        'category_id': request.args.get('category_id', type=int),
        'wallet_id': request.args.get('wallet_id', type=int),
        'start_date': request.args.get('start_date'),
        'end_date': request.args.get('end_date'),
        'search': request.args.get('search', ''),
    }

    transactions = get_filtered_transactions(current_user.id, filters).order_by(Transaction.date.desc()).limit(100).all()
    totals = calculate_transaction_totals(transactions)

    return render_template(
        'report_preview.html',
        transactions=transactions,
        filters=filters,
        total_income=totals['total_income'],
        total_expense=totals['total_expense'],
        net_total=totals['net_total'],
    )

@app.route('/transaction/add', methods=['POST'])
@login_required
def add_transaction():
    try:
        amount = parse_positive_amount(request.form.get('amount'))
    except ValueError as exc:
        flash(str(exc), 'danger')
        return redirect(url_for('transactions'))

    desc = request.form['description']
    ttype = request.form['type']

    try:
        cat_id = int(request.form['category_id'])
        wallet_id = int(request.form['wallet_id'])
        trans_date = parse_transaction_datetime(request.form.get('date'))
    except (TypeError, ValueError) as exc:
        flash(str(exc), 'danger')
        return redirect(url_for('transactions'))

    try:
        create_transaction(
            user_id=current_user.id,
            wallet_id=wallet_id,
            amount=amount,
            category_id=cat_id,
            description=desc,
            transaction_type=ttype,
            date=trans_date,
        )
        cache.delete_memoized(get_dashboard_data)
        logger.info(f"Transaction created: user={current_user.id}, amount={amount}")
        flash('Transaksi disimpan')
    except Exception as exc:
        logger.error('Transaction create error', exc_info=True)
        flash(str(exc), 'danger')
    return redirect(url_for('transactions'))

@app.route('/transaction/edit/<int:id>', methods=['POST'])
@login_required
def edit_transaction(id):
    trans = Transaction.query.get_or_404(id)
    try:
        trans_date = parse_transaction_datetime(request.form.get('date'))
        new_amount = parse_positive_amount(request.form.get('amount'))
        new_category_id = int(request.form['category_id'])
        new_wallet_id = int(request.form['wallet_id'])
        new_type = request.form['type']
    except (TypeError, ValueError) as exc:
        flash(str(exc), 'danger')
        return redirect(url_for('transactions'))
    
    try:
        update_transaction(
            transaction=trans,
            user_id=current_user.id,
            wallet_id=new_wallet_id,
            amount=new_amount,
            category_id=new_category_id,
            description=request.form['description'],
            transaction_type=new_type,
            date=trans_date,
        )
        cache.delete_memoized(get_dashboard_data)
        logger.info(f"Transaction updated: user={current_user.id}, transaction_id={id}, amount={new_amount}")
        flash('Transaksi diperbarui')
    except Exception as exc:
        logger.error('Transaction update error', exc_info=True)
        flash(str(exc), 'danger')
    return redirect(url_for('transactions'))

@app.route('/transaction/delete/<int:id>')
@login_required
def delete_transaction(id):
    trans = Transaction.query.get_or_404(id)
    try:
        delete_transaction_service(trans, current_user.id)
        cache.delete_memoized(get_dashboard_data)
        logger.info(f"Transaction deleted: user={current_user.id}, transaction_id={id}")
        flash('Transaksi dihapus')
    except Exception as exc:
        logger.error('Transaction delete error', exc_info=True)
        flash(str(exc), 'danger')
    return redirect(url_for('transactions'))

@app.route('/budgets')
@login_required
def budgets():
    from sqlalchemy import func, extract
    now = normalize_wib_storage(now_wib())
    month = request.args.get('month', now.month, type=int)
    year = request.args.get('year', now.year, type=int)
    budgets = Budget.query.filter_by(user_id=current_user.id, month=month, year=year).all()
    categories = Category.query.filter_by(user_id=current_user.id).all()
    
    # Hitung terpakai untuk setiap budget
    budget_data = []
    for budget in budgets:
        terpakai = db.session.query(func.sum(Transaction.amount)).\
            filter(Transaction.user_id == current_user.id,
                   Transaction.category_id == budget.category_id,
                   Transaction.type == 'expense',
                   extract('month', Transaction.date) == month,
                   extract('year', Transaction.date) == year).scalar() or 0
        budget_data.append({
            'budget': budget,
            'terpakai': float(terpakai)
        })
    
    return render_template('budgets.html', budget_data=budget_data, categories=categories, month=month, year=year)

@app.route('/budget/add', methods=['POST'])
@login_required
def add_budget():
    cat_id = int(request.form['category_id'])
    month = int(request.form['month'])
    year = int(request.form['year'])
    try:
        amount = parse_positive_amount(request.form.get('amount'))
    except ValueError as exc:
        flash(str(exc), 'danger')
        return redirect(url_for('budgets', month=month, year=year))
    # Cek apakah sudah ada
    existing = Budget.query.filter_by(user_id=current_user.id, category_id=cat_id, month=month, year=year).first()
    if existing:
        existing.amount = amount
    else:
        budget = Budget(category_id=cat_id, month=month, year=year, amount=amount, user_id=current_user.id)
        db.session.add(budget)
    db.session.commit()
    flash('Anggaran disimpan')
    return redirect(url_for('budgets', month=month, year=year))

@app.route('/budget/delete/<int:id>')
@login_required
def delete_budget(id):
    budget = Budget.query.get_or_404(id)
    if budget.user_id != current_user.id:
        flash('Anda tidak memiliki akses')
        return redirect(url_for('budgets'))
    month, year = budget.month, budget.year
    db.session.delete(budget)
    db.session.commit()
    flash('Anggaran dihapus')
    return redirect(url_for('budgets', month=month, year=year))

@app.route('/api/budget-details/<int:budget_id>')
@login_required
def budget_details(budget_id):
    budget = Budget.query.get_or_404(budget_id)
    if budget.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    
    from sqlalchemy import extract
    transactions = Transaction.query.\
        filter(Transaction.user_id == current_user.id,
               Transaction.category_id == budget.category_id,
               Transaction.type == 'expense',
               extract('month', Transaction.date) == budget.month,
               extract('year', Transaction.date) == budget.year).\
        order_by(Transaction.date.desc()).all()
    
    data = {
        'budget_category': budget.category.name,
        'budget_amount': budget.amount,
        'month': budget.month,
        'year': budget.year,
        'transactions': [{
            'date': to_wib(t.date).strftime('%d/%m/%Y'),
            'description': t.description,
            'amount': float(t.amount),
            'wallet': t.wallet.name
        } for t in transactions]
    }
    return jsonify(data)

@app.route('/export/budget-pdf/<int:budget_id>')
@login_required
def export_budget_pdf(budget_id):
    budget = Budget.query.get_or_404(budget_id)
    if budget.user_id != current_user.id:
        flash('Anda tidak memiliki akses')
        return redirect(url_for('budgets'))
    
    from sqlalchemy import extract, func
    transactions = Transaction.query.\
        filter(Transaction.user_id == current_user.id,
               Transaction.category_id == budget.category_id,
               Transaction.type == 'expense',
               extract('month', Transaction.date) == budget.month,
               extract('year', Transaction.date) == budget.year).\
        order_by(Transaction.date.asc()).all()
    
    total_spent = sum(t.amount for t in transactions)
    remaining = budget.amount - total_spent
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=portrait(letter), 
                            rightMargin=50, leftMargin=50, 
                            topMargin=50, bottomMargin=50)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=20,
        alignment=1, # Center
        textColor=colors.HexColor("#1e40af")
    )
    
    subtitle_style = ParagraphStyle(
        'SubtitleStyle',
        parent=styles['Normal'],
        fontSize=12,
        spaceAfter=12,
        alignment=1, # Center
        textColor=colors.grey
    )
    
    section_title_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceBefore=15,
        spaceAfter=10,
        textColor=colors.HexColor("#1e40af"),
        borderPadding=5,
        borderWidth=0,
        borderStyle=None
    )

    normal_style = styles['Normal']
    
    # Header
    elements.append(Paragraph("Laporan Detail Anggaran", title_style))
    elements.append(Paragraph(f"Periode: {budget.month}/{budget.year}", subtitle_style))
    elements.append(Spacer(1, 0.2 * inch))
    
    # Summary Info
    summary_data = [
        [Paragraph("<b>Kategori</b>", normal_style), f": {budget.category.name}"],
        [Paragraph("<b>Target Anggaran</b>", normal_style), f": Rp {budget.amount:,.0f}"],
        [Paragraph("<b>Total Terpakai</b>", normal_style), f": Rp {total_spent:,.0f}"],
        [Paragraph("<b>Sisa Anggaran</b>", normal_style), f": Rp {remaining:,.0f}"],
        [Paragraph("<b>Status</b>", normal_style), f": {'Melebihi' if total_spent > budget.amount else 'Aman'}"]
    ]
    
    summary_table = Table(summary_data, colWidths=[1.5*inch, 4*inch])
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TEXTCOLOR', (1,2), (1,2), colors.red if total_spent > budget.amount else colors.green),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 11),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3 * inch))
    
    # Transactions Table
    elements.append(Paragraph("Rincian Transaksi", section_title_style))
    
    table_data = [['Tanggal', 'Deskripsi', 'Dompet', 'Jumlah']]
    for t in transactions:
        table_data.append([
            to_wib(t.date).strftime('%d/%m/%Y'),
            Paragraph(t.description, normal_style),
            t.wallet.name,
            f"Rp {t.amount:,.0f}"
        ])
    
    if len(transactions) == 0:
        table_data.append(['-', 'Tidak ada transaksi', '-', '-'])

    t_table = Table(table_data, colWidths=[1.0*inch, 2.5*inch, 1.2*inch, 1.3*inch])
    t_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1e40af")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('ALIGN', (3,1), (3,-1), 'RIGHT'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 10),
        ('TOPPADDING', (0,0), (-1,0), 10),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.whitesmoke, colors.white])
    ]))
    elements.append(t_table)
    
    # Footer
    elements.append(Spacer(1, 0.5 * inch))
    footer_text = f"Dicetak pada: {now_wib().strftime('%d/%m/%Y %H:%M:%S')}"
    elements.append(Paragraph(footer_text, ParagraphStyle('Footer', parent=styles['Italic'], alignment=2, fontSize=8)))

    doc.build(elements)
    buffer.seek(0)
    filename = f"Laporan_Anggaran_{budget.category.name}_{budget.month}_{budget.year}.pdf"
    return send_file(buffer, download_name=filename, as_attachment=True)

@app.route('/reports')
@login_required
def reports():
    return render_template('reports.html')

@app.route('/api/chart-data')
@login_required
def chart_data():
    from sqlalchemy import func, extract
    now = normalize_wib_storage(now_wib())
    data = db.session.query(Category.name, func.sum(Transaction.amount)).\
           join(Transaction).\
           filter(Transaction.user_id == current_user.id,
                  Transaction.type == 'expense',
                  extract('month', Transaction.date) == now.month,
                  extract('year', Transaction.date) == now.year).\
           group_by(Category.id).all()
    return jsonify({'labels': [d[0] for d in data], 'values': [float(d[1]) for d in data]})

@app.route('/api/income-expense-data')
@login_required
def income_expense_data():
    from sqlalchemy import func, extract
    now = normalize_wib_storage(now_wib())
    
    # Get monthly income and expense totals
    income = db.session.query(func.sum(Transaction.amount)).\
             filter(Transaction.user_id == current_user.id,
                    Transaction.type == 'income',
                    extract('month', Transaction.date) == now.month,
                    extract('year', Transaction.date) == now.year).scalar() or 0
    
    expense = db.session.query(func.sum(Transaction.amount)).\
              filter(Transaction.user_id == current_user.id,
                     Transaction.type == 'expense',
                     extract('month', Transaction.date) == now.month,
                     extract('year', Transaction.date) == now.year).scalar() or 0
    
    return jsonify({
        'labels': ['Pemasukan', 'Pengeluaran'],
        'income': float(income),
        'expense': float(expense)
    })

@app.route('/api/income-expense-line')
@login_required
def income_expense_line():
    from sqlalchemy import func, extract
    now = normalize_wib_storage(now_wib())
    labels = []
    incomes = []
    expenses = []
    # last 6 months
    for i in range(5, -1, -1):
        year = now.year
        month = now.month - i
        while month <= 0:
            month += 12
            year -= 1
        label = f"{month}/{year}"
        inc = db.session.query(func.sum(Transaction.amount)).filter(
            Transaction.user_id == current_user.id,
            Transaction.type == 'income',
            extract('month', Transaction.date) == month,
            extract('year', Transaction.date) == year
        ).scalar() or 0
        exp = db.session.query(func.sum(Transaction.amount)).filter(
            Transaction.user_id == current_user.id,
            Transaction.type == 'expense',
            extract('month', Transaction.date) == month,
            extract('year', Transaction.date) == year
        ).scalar() or 0
        labels.append(label)
        incomes.append(float(inc))
        expenses.append(float(exp))
    return jsonify({'labels': labels, 'income': incomes, 'expense': expenses})

@app.route('/api/budget-realization')
@login_required
def budget_realization():
    from sqlalchemy import func, extract
    now = normalize_wib_storage(now_wib())
    budgets = Budget.query.filter_by(user_id=current_user.id, month=now.month, year=now.year).all()
    labels = []
    budget_vals = []
    real_vals = []
    for b in budgets:
        labels.append(b.category.name)
        budget_vals.append(b.amount)
        real = db.session.query(func.sum(Transaction.amount)).filter(
            Transaction.user_id == current_user.id,
            Transaction.category_id == b.category_id,
            Transaction.type == 'expense',
            extract('month', Transaction.date) == now.month,
            extract('year', Transaction.date) == now.year
        ).scalar() or 0
        real_vals.append(float(real))
    return jsonify({'labels': labels, 'budget': budget_vals, 'real': real_vals})

@app.route('/api/cashflow-data')
@login_required
def cashflow_data():
    now = normalize_wib_storage(now_wib())
    start = now - timedelta(days=30)
    transactions = Transaction.query.filter(Transaction.user_id == current_user.id,
                                         Transaction.date >= start).order_by(Transaction.date).all()
    labels = []
    balances = []
    bal = 0
    for t in transactions:
        if t.type == 'income':
            bal += t.amount
        else:
            bal -= t.amount
        labels.append(to_wib(t.date).strftime('%d/%m'))
        balances.append(bal)
    return jsonify({'labels': labels, 'balance': balances})

@app.route('/export/excel')
@login_required
def export_excel():
    filters = {
        'category_id': request.args.get('category_id', type=int),
        'wallet_id': request.args.get('wallet_id', type=int),
        'start_date': request.args.get('start_date'),
        'end_date': request.args.get('end_date'),
        'search': request.args.get('search', ''),
    }
    transactions = get_filtered_transactions(current_user.id, filters).order_by(Transaction.date.desc()).all()
    
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = 'Transaksi'
    
    # Define header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add headers
    headers = ['Tanggal', 'Deskripsi', 'Jumlah', 'Tipe', 'Kategori', 'Dompet']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add data rows
    for row_num, transaction in enumerate(transactions, 2):
        ws.cell(row=row_num, column=1).value = to_wib(transaction.date).strftime('%Y-%m-%d %H:%M')
        ws.cell(row=row_num, column=2).value = transaction.description
        ws.cell(row=row_num, column=3).value = transaction.amount
        ws.cell(row=row_num, column=4).value = 'Pemasukan' if transaction.type == 'income' else 'Pengeluaran'
        ws.cell(row=row_num, column=5).value = transaction.category.name
        ws.cell(row=row_num, column=6).value = transaction.wallet.name
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='laporan_keuangan.xlsx', as_attachment=True)

@app.route('/export/pdf')
@login_required
def export_pdf():
    filters = {
        'category_id': request.args.get('category_id', type=int),
        'wallet_id': request.args.get('wallet_id', type=int),
        'start_date': request.args.get('start_date'),
        'end_date': request.args.get('end_date'),
        'search': request.args.get('search', ''),
    }
    start_date = filters['start_date']
    end_date = filters['end_date']

    transactions = get_filtered_transactions(current_user.id, filters).order_by(Transaction.date.asc()).all()
    
    # Calculate summary
    totals = calculate_transaction_totals(transactions)
    total_income = totals['total_income']
    total_expense = totals['total_expense']
    net_flow = totals['net_total']

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=portrait(letter),
                            rightMargin=40, leftMargin=40,
                            topMargin=40, bottomMargin=40)
    elements = []
    
    styles = getSampleStyleSheet()
    
    # Custom Styles
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontSize=20,
        spaceAfter=10,
        alignment=1,  # Center
        textColor=colors.HexColor("#1e40af")
    )
    
    subtitle_style = ParagraphStyle(
        'SubtitleStyle',
        parent=styles['Normal'],
        fontSize=11,
        spaceAfter=20,
        alignment=1,  # Center
        textColor=colors.grey
    )

    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceBefore=15,
        spaceAfter=10,
        textColor=colors.HexColor("#1e40af")
    )

    normal_style = styles['Normal']
    
    # 1. Title & Header
    elements.append(Paragraph("LAPORAN KEUANGAN", title_style))
    date_range = "Semua Waktu"
    if start_date and end_date:
        date_range = f"{start_date} s/d {end_date}"
    elif start_date:
        date_range = f"Mulai {start_date}"
    elif end_date:
        date_range = f"Hingga {end_date}"
        
    elements.append(Paragraph(f"Periode: {date_range}", subtitle_style))
    elements.append(Paragraph(f"User: {current_user.username}", subtitle_style))
    elements.append(Spacer(1, 0.2 * inch))

    # 2. Summary Card
    elements.append(Paragraph("Ringkasan", header_style))
    summary_data = [
        ["Total Pemasukan", f"Rp {total_income:,.0f}"],
        ["Total Pengeluaran", f"Rp {total_expense:,.0f}"],
        ["Arus Kas Bersih", f"Rp {net_flow:,.0f}"]
    ]
    
    summary_table = Table(summary_data, colWidths=[2.5*inch, 2.5*inch])
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (0,-1), 'LEFT'),
        ('ALIGN', (1,0), (1,-1), 'RIGHT'),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 11),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.lightgrey),
        ('TEXTCOLOR', (1,2), (1,2), colors.green if net_flow >= 0 else colors.red),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.4 * inch))

    # 3. Transactions Table
    elements.append(Paragraph("Rincian Transaksi", header_style))
    
    table_data = [['Tanggal', 'Deskripsi', 'Kategori', 'Tipe', 'Jumlah']]
    for t in transactions:
        table_data.append([
            to_wib(t.date).strftime('%d/%m/%Y'),
            Paragraph(t.description, normal_style),
            t.category.name,
            'Masuk' if t.type == 'income' else 'Keluar',
            f"Rp {t.amount:,.0f}"
        ])
    
    if not transactions:
        table_data.append(['-', 'Tidak ada transaksi ditemukan', '-', '-', '-'])

    t_table = Table(table_data, colWidths=[1.0*inch, 2.4*inch, 1.2*inch, 0.8*inch, 1.5*inch])
    t_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1e40af")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('ALIGN', (4,1), (4,-1), 'RIGHT'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 10),
        ('TOPPADDING', (0,0), (-1,0), 10),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.whitesmoke, colors.white]),
        ('TEXTCOLOR', (3,1), (3,-1), colors.green), # Default column 3 color (Type)
    ]))
    
    # Conditional coloring for type column
    if transactions:
        for i, t in enumerate(transactions, 1):
            if t.type == 'expense':
                t_table.setStyle(TableStyle([
                    ('TEXTCOLOR', (3, i), (3, i), colors.red),
                    ('TEXTCOLOR', (4, i), (4, i), colors.red)
                ]))
            else:
                t_table.setStyle(TableStyle([
                    ('TEXTCOLOR', (3, i), (3, i), colors.green),
                    ('TEXTCOLOR', (4, i), (4, i), colors.green)
                ]))

    elements.append(t_table)
    
    # 4. Footer
    elements.append(Spacer(1, 0.5 * inch))
    footer_text = f"Dicetak pada: {now_wib().strftime('%d/%m/%Y %H:%M:%S')}"
    elements.append(Paragraph(footer_text, ParagraphStyle('Footer', parent=styles['Italic'], alignment=2, fontSize=8, textColor=colors.grey)))

    doc.build(elements)
    buffer.seek(0)
    
    filename = f"Laporan_Keuangan_{current_user.username}_{now_wib().strftime('%Y%m%d')}.pdf"
    return send_file(buffer, download_name=filename, as_attachment=True)

@app.route('/share/wallet', methods=['POST'])
@login_required
def share_wallet():
    wallet_id = request.form['wallet_id']
    username = request.form['username']
    permission = request.form['permission']
    user_to_share = User.query.filter_by(username=username).first()
    if not user_to_share:
        flash('User tidak ditemukan')
        return redirect(url_for('wallets'))
    # Cek apakah wallet milik current_user
    wallet = Wallet.query.get(wallet_id)
    if wallet.user_id != current_user.id:
        flash('Anda hanya dapat membagikan dompet milik sendiri')
        return redirect(url_for('wallets'))
    # Cek apakah sudah ada
    existing = SharedWallet.query.filter_by(wallet_id=wallet_id, shared_with_id=user_to_share.id).first()
    if existing:
        existing.permission = permission
    else:
        share = SharedWallet(wallet_id=wallet_id, shared_with_id=user_to_share.id, permission=permission)
        db.session.add(share)
    db.session.commit()
    flash('Dompet dibagikan')
    return redirect(url_for('wallets'))

@app.route('/unshare/wallet/<int:id>')
@login_required
def unshare_wallet(id):
    share = SharedWallet.query.get_or_404(id)
    # Pastikan yang menghapus adalah pemilik dompet
    if share.wallet.user_id != current_user.id:
        flash('Anda tidak memiliki izin')
        return redirect(url_for('wallets'))
    db.session.delete(share)
    db.session.commit()
    flash('Berbagi dompet dihentikan')
    return redirect(url_for('wallets'))

# ================== FITUR TRANSFER ==================
@app.route('/transfer', methods=['GET', 'POST'])
@login_required
def transfer():
    wallets = Wallet.query.filter_by(user_id=current_user.id).all()
    if request.method == 'POST':
        try:
            from_wallet_id = int(request.form['from_wallet'])
            to_wallet_id = int(request.form['to_wallet'])
            amount = parse_positive_amount(request.form.get('amount'))
            fee = parse_positive_amount(request.form.get('fee', 0), 'Biaya transfer', allow_zero=True)
        except (TypeError, ValueError) as exc:
            flash(str(exc), 'danger')
            return redirect(url_for('transfer'))

        description = request.form.get('description', 'Transfer')

        try:
            transfer_balance(
                user_id=current_user.id,
                from_wallet_id=from_wallet_id,
                to_wallet_id=to_wallet_id,
                amount=amount,
                fee=fee,
                description=description,
            )
            cache.delete_memoized(get_dashboard_data)
            logger.info(
                f"Transfer completed: user={current_user.id}, from_wallet={from_wallet_id}, to_wallet={to_wallet_id}, amount={amount}, fee={fee}"
            )
            flash('Transfer berhasil')
        except Exception as exc:
            logger.error('Transfer failed', exc_info=True)
            flash(str(exc), 'danger')
            return redirect(url_for('transfer'))

        return redirect(url_for('wallets'))

    return render_template('transfer.html', wallets=wallets)

@app.route('/profile')
@login_required
def profile():
    return render_template('profile.html')

@app.route('/upload_photo', methods=['POST'])
@login_required
def upload_photo():
    if 'photo' not in request.files:
        flash('Tidak ada file yang dipilih')
        return redirect(url_for('profile'))
    
    file = request.files['photo']
    if file.filename == '':
        flash('Tidak ada file yang dipilih')
        return redirect(url_for('profile'))

    if not allowed_file(file.filename):
        flash('Format file tidak didukung. Gunakan JPG, PNG, GIF, atau WEBP')
        return redirect(url_for('profile'))

    try:
        new_relative_path = optimize_and_save_profile_image(file, current_user.id)

        # Hapus foto lama jika ada dan berbeda dengan file baru
        if current_user.photo:
            old_photo_path = safe_static_file_path(current_user.photo)
            new_photo_path = safe_static_file_path(new_relative_path)
            if old_photo_path and os.path.exists(old_photo_path) and old_photo_path != new_photo_path:
                os.remove(old_photo_path)

        # Update database
        current_user.photo = new_relative_path
        db.session.commit()

        flash('Foto profil berhasil diubah')
    except ValueError as e:
        flash(str(e))
    except Exception:
        flash('Gagal memproses foto. Silakan coba gambar lain.')

    return redirect(url_for('profile'))

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_IMAGE_EXTENSIONS


def safe_static_file_path(relative_path):
    if not relative_path:
        return None

    normalized = relative_path.replace('\\', '/').lstrip('/')
    if normalized.startswith('static/'):
        normalized = normalized[len('static/'):]

    static_root = os.path.abspath(os.path.join(app.root_path, 'static'))
    full_path = os.path.abspath(os.path.join(static_root, normalized))
    if not full_path.startswith(static_root + os.sep) and full_path != static_root:
        return None
    return full_path


def get_profile_photo_url(relative_path):
    file_path = safe_static_file_path(relative_path)
    if not file_path or not os.path.exists(file_path):
        return None

    normalized = relative_path.replace('\\', '/').lstrip('/')
    if normalized.startswith('static/'):
        normalized = normalized[len('static/'):]
    return url_for('static', filename=normalized)


def image_has_transparency(image):
    if image.mode in ('RGBA', 'LA'):
        alpha = image.getchannel('A')
        return alpha.getextrema()[0] < 255
    if image.mode == 'P':
        return 'transparency' in image.info
    return False


def optimize_and_save_profile_image(file_storage, user_id):
    try:
        file_storage.stream.seek(0)
        image = Image.open(file_storage.stream)
        image.verify()
    except (UnidentifiedImageError, OSError):
        raise ValueError('File bukan gambar yang valid')

    try:
        file_storage.stream.seek(0)
        image = Image.open(file_storage.stream)
        image = ImageOps.exif_transpose(image)
    except Exception:
        raise ValueError('Gagal membaca data gambar')

    width, height = image.size
    if width <= 0 or height <= 0:
        raise ValueError('Ukuran gambar tidak valid')
    if width * height > MAX_PROFILE_IMAGE_PIXELS:
        raise ValueError('Resolusi gambar terlalu besar')

    if max(width, height) > MAX_PROFILE_IMAGE_DIMENSION:
        image.thumbnail((MAX_PROFILE_IMAGE_DIMENSION, MAX_PROFILE_IMAGE_DIMENSION), Image.Resampling.LANCZOS)

    original_ext = file_storage.filename.rsplit('.', 1)[1].lower()
    keep_png = (original_ext == 'png' and image_has_transparency(image))

    if keep_png:
        filename = f"user_{user_id}.png"
        output_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        temp_output = f"{output_path}.tmp"

        if image.mode != 'RGBA':
            image = image.convert('RGBA')

        image.save(temp_output, format='PNG', optimize=True)
        os.replace(temp_output, output_path)
        return f'uploads/{filename}'

    filename = f"user_{user_id}.jpg"
    output_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    temp_output = f"{output_path}.tmp"

    if image.mode not in ('RGB', 'L'):
        image = image.convert('RGB')
    elif image.mode == 'L':
        image = image.convert('RGB')

    image.save(
        temp_output,
        format='JPEG',
        quality=PROFILE_JPEG_QUALITY,
        optimize=True,
        progressive=True,
        subsampling='4:2:0'
    )
    os.replace(temp_output, output_path)
    return f'uploads/{filename}'

@app.route('/change_password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        current_password = request.form['current_password']
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']
        
        if not current_user.check_password(current_password):
            flash('Password saat ini salah')
            return redirect(url_for('change_password'))
        
        if new_password != confirm_password:
            flash('Password baru tidak cocok')
            return redirect(url_for('change_password'))
        
        current_user.set_password(new_password)  # Hash password baru sebelum menyimpan
        db.session.commit()
        flash('Password berhasil diubah')
        return redirect(url_for('profile'))
    
    return render_template('change_password.html')

@app.route('/backup')
@login_required
def backup():
    # Backup data ke CSV
    import csv
    from io import StringIO
    
    output = StringIO()
    writer = csv.writer(output)
    
    # Header untuk transactions
    writer.writerow(['Date', 'Amount', 'Description', 'Type', 'Category', 'Wallet'])
    
    transactions = Transaction.query.filter_by(user_id=current_user.id).all()
    for t in transactions:
        category = Category.query.get(t.category_id)
        wallet = Wallet.query.get(t.wallet_id)
        writer.writerow([
            to_wib(t.date).strftime('%Y-%m-%d %H:%M:%S'),
            t.amount,
            t.description,
            t.type,
            category.name if category else '',
            wallet.name if wallet else ''
        ])
    
    output.seek(0)
    return send_file(
        BytesIO(output.getvalue().encode('utf-8')),
        mimetype='text/csv',
        as_attachment=True,
        download_name='finance_backup.csv'
    )

@app.route('/import_data', methods=['GET', 'POST'])
@login_required
def import_data():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Tidak ada file yang dipilih')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('Tidak ada file yang dipilih')
            return redirect(request.url)
        
        if file and file.filename.endswith('.csv'):
            import csv
            from io import TextIOWrapper
            
            stream = TextIOWrapper(file.stream, encoding='utf-8')
            csv_reader = csv.reader(stream)
            
            # Skip header
            next(csv_reader, None)
            
            imported_count = 0
            errors = []
            
            for row_num, row in enumerate(csv_reader, start=2):
                try:
                    if len(row) != 6:
                        errors.append(f'Baris {row_num}: Format tidak valid')
                        continue
                    
                    date_str, amount_str, description, type_, category_name, wallet_name = row
                    
                    # Parse date
                    try:
                        date = normalize_wib_storage(datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S'))
                    except ValueError:
                        errors.append(f'Baris {row_num}: Format tanggal tidak valid')
                        continue
                    
                    # Parse amount
                    try:
                        amount = parse_positive_amount(amount_str)
                    except ValueError:
                        errors.append(f'Baris {row_num}: Jumlah tidak valid')
                        continue
                    
                    # Validate type
                    if type_ not in ['income', 'expense']:
                        errors.append(f'Baris {row_num}: Tipe harus income atau expense')
                        continue
                    
                    # Get or create category
                    category = Category.query.filter_by(user_id=current_user.id, name=category_name, type=type_).first()
                    if not category:
                        category = Category(name=category_name, type=type_, user_id=current_user.id)
                        db.session.add(category)
                        db.session.flush()
                    
                    # Get or create wallet
                    wallet = Wallet.query.filter_by(user_id=current_user.id, name=wallet_name).first()
                    if not wallet:
                        wallet = Wallet(name=wallet_name, type='cash', balance=0.0, user_id=current_user.id)
                        db.session.add(wallet)
                        db.session.flush()
                    
                    # Create transaction
                    transaction = Transaction(
                        amount=amount,
                        description=description,
                        date=date,
                        type=type_,
                        category_id=category.id,
                        wallet_id=wallet.id,
                        user_id=current_user.id
                    )
                    db.session.add(transaction)
                    
                    # Update wallet balance
                    if type_ == 'income':
                        wallet.balance += amount
                    else:
                        wallet.balance -= amount
                    
                    imported_count += 1
                    
                except Exception as e:
                    errors.append(f'Baris {row_num}: Error - {str(e)}')
            
            db.session.commit()
            
            if imported_count > 0:
                flash(f'Berhasil mengimpor {imported_count} transaksi')
            if errors:
                flash(f'Error pada {len(errors)} baris: ' + '; '.join(errors[:5]))  # Show first 5 errors
            
            return redirect(url_for('profile'))
        else:
            flash('File harus berformat CSV')
            return redirect(request.url)
    
    return render_template('import_data.html')
